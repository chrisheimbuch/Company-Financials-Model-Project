# %%
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# %%
df = pd.read_csv("Financials.csv")
df.head()


# %%
df.shape

# %%
df.info()

# %%
df.describe()

# %%
df.head()

# %%
df_copy = df.copy()

# %%
df_copy.columns = df.columns.str.lower()

# %%
df_copy.columns

# %%
cleaned_columns = [col.strip() for col in df_copy.columns]

# %%
cleaned_columns

# %%
df_copy.columns = cleaned_columns

# %%
df_copy.columns

# %%
rows_to_adjust = ["units sold", "manufacturing price", "sale price", "gross sales", "discounts", "sales", "cogs", "profit"]

# %%
df_copy

# %%
print(df_copy['gross sales'].head(10))

print(df_copy['gross sales'].dtype)

# %%
#Adjust the rows with dollar signs to remove them.
df_copy[rows_to_adjust] = df_copy[rows_to_adjust].apply(lambda x: x.str.replace('$', ""))

# %%
#Adjust the rows with commas to remove them.
df_copy[rows_to_adjust] = df_copy[rows_to_adjust].apply(lambda x: x.str.replace(',', ""))

# %%
#Adjust the rows with commas to remove them.
df_copy[rows_to_adjust] = df_copy[rows_to_adjust].apply(lambda x: x.str.strip())

# %%
#Adjust leading and trailing white space for all columns
additional_adjustment_columns = ['segment', 'country', 'product', 'discount band', 'month name']
df_copy[additional_adjustment_columns] = df_copy[additional_adjustment_columns].apply(lambda x: x.str.strip())

# %%
#Adjust the rows with dashes to remove them and replace with 0's.
df_copy['discounts'] = df_copy['discounts'].str.replace("-", "0")

# %%
df_copy

# %%
#It still seems there are some problems with "-". Let's try to see which columns have it.

for col in rows_to_adjust:
    print(f"Column: {col}")
    print("Contains '-':", (df_copy[col] == "-").sum()) 
    print("Sample values:", df_copy[col].head(10).tolist())
    print("---")

# %%
#this confirms that there are 5 instance of only "-" within profit and therefore we can replace them with 0's.
df_copy[df_copy['profit'] == "-"]

# %%
df_copy['profit'] = df_copy['profit'].str.replace("-", "0")

# %%
#run the same code and no longer exist - perfect.
df_copy[df_copy['profit'] == "-"]


# %%
#See which rows contain negative values

for col in rows_to_adjust:
    print(f"Column: {col}")
    print("Contains parenthesis:", (df_copy[col].str.contains(r"\(.*\)", na=False)).sum())
    print("Sample values:", df_copy[col].head(10).tolist())
    print("---")

# %%
df_copy['profit'] = df_copy['profit'].str.replace("(", "-")
df_copy['profit'] = df_copy['profit'].str.replace(")", "")

# %%
for col in rows_to_adjust:
    print(f"Column: {col}")
    print("Contains parenthesis:", (df_copy[col].str.contains(r"\(.*\)", na=False)).sum())
    print("Sample values:", df_copy[col].head(10).tolist())
    print("---")

# %%
df_copy[rows_to_adjust] = df_copy[rows_to_adjust].astype(float)

# %%
df_copy.info()

# %%
df_copy

# %% [markdown]
# ## Now that the dataframe is clean of leading and trailing 0s, dollar signs, erroneous commas, and in a format that is workable to extract meaningful information to questions, we can start working with it within this workbook, and also get it in a format where it is excel ready.

# %%
# Export DataFrame to Excel 
output_path = "C:/Users/Chris/Desktop/Documents/Company Financials Model Project/cleaned_financials.xlsx"
df_copy.to_excel(output_path, sheet_name="Financials", index=False)
print("DataFrame exported to Excel!")

# %%
# How many unique years are there that we are working with in this data set? Answer: 2 - 2013 and 2014
df_copy['year'].unique()

# %%
# Which year had better profitability?

profit_by_year = df_copy.groupby("year")["profit"].sum().reset_index()
profit_by_year.sort_values(by="profit", ascending=False)
profit_by_year['year']= profit_by_year['year'].astype(str)

# Plot the bar chart
fig, ax = plt.subplots(figsize=(12,8))
ax.bar(x=profit_by_year['year'], height=profit_by_year['profit'], color='#B85042', edgecolor='black', zorder=2)

#Custom Font Dictionaries
font1 = {'family':'serif','color':'black','size':16}
font2 = {'family':'serif','color':'black','size':14}

#Set title, axis names, customize fonts.
ax.set_xlabel("Segment", fontdict=font1)
ax.set_ylabel("Profit", fontdict=font2)
ax.set_title("Profit by Segment", fontdict=font1)

#Customize graph
ax.grid(axis = 'y', color='grey')
ax.set_facecolor('#E7E8D1')
plt.xticks(rotation=0, ha='center', fontsize=13)
plt.yticks(fontsize=13)

ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))

#Annotate each bar
for bar in ax.patches:
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width() / 2, 
        height - 50000, 
        f'{height: ,.0f}', 
        ha='center', 
        va='top', 
        color='black', 
        fontsize=12)


plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_year.png", dpi=100, bbox_inches="tight")

# %%
# Embed graph in Excel
output_path = "C:/Users/Chris/Desktop/Documents/Company Financials Model Project/cleaned_financials.xlsx"
book = load_workbook(output_path)
sheet = book.create_sheet("Charts")
img = Image("Graph_Visuals/profit_by_year.png")
sheet.add_image(img, "A20")  # Adjust cell position
book.save(output_path)
print("Graph embedded in Excel!")

# %%
#Create a sample graph (e.g., profit by company)

plt.figure(figsize=(10, 6))
# Replace with your column name (e.g., "Company")
profit_by_segment = df_copy.groupby("segment")["profit"].sum().reset_index()
profit_by_segment.sort_values(by="profit", ascending=False)

# Plot the bar chart
fig, ax = plt.subplots(figsize=(12,8))
ax.bar(x=profit_by_segment['segment'], height=profit_by_segment['profit'], color='#B85042', edgecolor='black', zorder=2)

#Custom Font Dictionaries
font1 = {'family':'serif','color':'black','size':16}
font2 = {'family':'serif','color':'black','size':14}

#Set title, axis names, customize fonts.
ax.set_xlabel("Segment", fontdict=font1)
ax.set_ylabel("Profit", fontdict=font2)
ax.set_title("Profit by Segment", fontdict=font1)

#Customize graph
ax.grid(axis = 'y', color='grey')
ax.set_facecolor('#E7E8D1')
plt.xticks(rotation=45, ha='center', fontsize=13)
plt.yticks(fontsize=13)

ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))

#Annotate each bar
for bar in ax.patches:
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width() / 2, 
        height - 50000, 
        f'{height: ,.0f}', 
        ha='center', 
        va='top', 
        color='black', 
        fontsize=12)


plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_segment.png", dpi=100, bbox_inches="tight")



# %% [markdown]
# ### Enterprise is a loss overall for the company of ~614k. I would like to understand why this is happening. There could be a high discount band for most of these customers. If this proves to be true, perhaps we could understand the reason for these discount bands. 

# %%
#Let's make a mask of the dataframe for segment and year to see what is happening.

MASKED_SEGMENT = df_copy['segment'] == "Enterprise"
enterprise_df = df_copy[MASKED_SEGMENT]
MASKED_SEGMENT_AND_YEAR = enterprise_df['year'] == 2014
enterprise_df_adjusted = enterprise_df[MASKED_SEGMENT_AND_YEAR]
enterprise_df_adjusted



# %%
counts = enterprise_df_adjusted['discount band'].value_counts()

def plot_bar(x_data, y_data, plot_title, x_name, y_name, bar_color='#CC313D', face_color='#F7C5CC', annotate_color='black', annotate_font=10, annotate_placement=1):

    #Custom Fonts
    font1 = {'family':'verdana','color':'#000000','size':20}
    font2 = {'family':'verdana','color':'#000000','size':16}

    #Create the plot, set x & y axis titles, and graph title.
    fig, ax = plt.subplots(figsize=(12,8))
    ax.bar(x=x_data, height=y_data, color=bar_color, edgecolor='black', zorder=3)
    ax.set_title(plot_title, fontdict=font1)
    ax.set_xlabel(x_name, fontdict=font2)
    ax.set_ylabel(y_name,fontdict=font2)

    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["top"].set_visible(False)

    #Plot Styling for axes ticks
    plt.xticks(fontsize=14)
    plt.yticks(fontsize=14)
    ax.set_facecolor(face_color)



    for bar in ax.patches:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2, 
            height + annotate_placement, 
            f'{height:.0f}', 
            ha='center', 
            va='top', 
            color=annotate_color, 
            fontsize=annotate_font)

    ax.grid(axis='y')
    plt.xticks(ha='center')

    plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/enterprise_discount_frequency.png", dpi=100, bbox_inches="tight")
  

plot_bar(counts.index, counts.values, "Discount Band for Enterprise Frequency", "Discount Type", "Frequency", '#2F3C7E', '#FBEAEB', 'white', 14, -1)


     

# %%


# %%
# Embed graph in Excel
output_path = "C:/Users/Chris/Desktop/Documents/Company Financials Model Project/cleaned_financials.xlsx"
book = load_workbook(output_path)
sheet = book.create_sheet("Charts")
img = Image("Graph_Visuals/profit_by_segment.png")
sheet.add_image(img, "A40")  # Adjust cell position
book.save(output_path)
print("Graph embedded in Excel!")

# %%
#Create a sample graph (e.g., profit by company)

plt.figure(figsize=(10, 6))
# Replace with your column name (e.g., "Company")
profit_by_product = df_copy.groupby("product")["profit"].sum().reset_index()
profit_by_product.sort_values(by="profit", ascending=False)

# Plot the bar chart
fig, ax = plt.subplots(figsize=(12,8))
ax.bar(x=profit_by_product['product'], height=profit_by_product['profit'], color='#B85042', edgecolor='black', zorder=2)

#Custom Font Dictionaries
font1 = {'family':'serif','color':'black','size':16}
font2 = {'family':'serif','color':'black','size':14}

#Set title, axis names, customize fonts.
ax.set_xlabel("Product type", fontdict=font1)
ax.set_ylabel("Profit", fontdict=font2)
ax.set_title("Profit by Product", fontdict=font1)

#Customize graph
ax.grid(axis = 'y', color='grey')
ax.set_facecolor('#E7E8D1')
plt.xticks(rotation=0, ha='center', fontsize=13)
plt.yticks(fontsize=13)

ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))

#Annotate each bar
for bar in ax.patches:
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width() / 2, 
        height - 50000, 
        f'{height: ,.0f}', 
        ha='center', 
        va='top', 
        color='black', 
        fontsize=12)


plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_product.png", dpi=100, bbox_inches="tight")


# %%
# Embed graph in Excel
output_path = "C:/Users/Chris/Desktop/Documents/Company Financials Model Project/cleaned_financials.xlsx"
book = load_workbook(output_path)
sheet = book.create_sheet("Charts")
img = Image("Graph_Visuals/profit_by_product.png")
sheet.add_image(img, "A60")  # Adjust cell position
book.save(output_path)
print("Graph embedded in Excel!")

# %%
profit_by_country = df_copy.groupby("country")["profit"].sum().reset_index()
profit_by_country.sort_values(by="profit", ascending=False)

fig, ax = plt.subplots(figsize=(12, 8))
ax.bar(x=profit_by_country['country'], height=profit_by_country['profit'], color='#B85042', edgecolor='black', zorder=3)
ax.set_xlabel("Country", fontdict={'family':'serif','color':'black','size':16})
ax.set_ylabel("Profit (USD)", fontdict={'family':'serif','color':'black','size':14})
ax.set_title("Profit by Country", fontdict={'family':'serif','color':'black','size':16})
ax.grid(axis='y', color='grey')
ax.set_facecolor('#E7E8D1')
plt.xticks(rotation=45, ha='center', fontsize=13)
plt.yticks(fontsize=13)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))
for bar in ax.patches:
    height = bar.get_height()
    label_pos = height + 50000 if height < 0 else height - 50000
    va_pos = 'bottom' if height < 0 else 'top'
    ax.text(bar.get_x() + bar.get_width() / 2, 
            label_pos, 
            f'{height:,.0f}', 
            ha='center', 
            va=va_pos, 
            color='black', 
            fontsize=12)
plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_country.png", dpi=100, bbox_inches="tight")


# %%
#This confirms that months are listed alphabetically and not chronologically which would be better for visualization.

# %%
#Create mask for 2014 only
YEAR_MASK = df_copy['year'] == 2014

#Apply mask to new df.
df_2014 = df_copy[YEAR_MASK]

# %%
# Question: Which month has the best profit in 2014?

profit_by_month = df_2014.groupby("month name")["profit"].sum().reset_index()

#Create chronological order for months instead of alphabetical order.
month_order = ["January", "February", "March", "April", "May", "June", 
               "July", "August", "September", "October", "November", "December"]

#Convert month to categorical with specified values.
profit_by_month["month name"] = pd.Categorical(profit_by_month["month name"], 
                                               categories=month_order, 
                                               ordered=True)
#Sort by the categorical order
profit_by_month = profit_by_month.sort_values("month name")

#Set the plot for the graph. Set the name of the X and Y axis, and title of the graph. Additionally, set some styling for aesthetic.
fig, ax = plt.subplots(figsize=(12, 8))
ax.bar(x=profit_by_month['month name'], height=profit_by_month['profit'], color='#B85042', edgecolor='black', zorder=3)
ax.set_xlabel("Month", fontdict={'family':'serif','color':'black','size':16})
ax.set_ylabel("Profit (USD)", fontdict={'family':'serif','color':'black','size':14})
ax.set_title("Profit by Month", fontdict={'family':'serif','color':'black','size':16})

#Plot Styling
ax.grid(axis='y', color='grey')
ax.set_facecolor('#E7E8D1')
plt.xticks(rotation=45, ha='center', fontsize=13)
plt.yticks(fontsize=13)
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{int(x):,}"))

#for loop to annotate each bar.
for bar in ax.patches:
    height = bar.get_height()
    label_pos = height + 50000 if height < 0 else height - 50000
    va_pos = 'bottom' if height < 0 else 'top'
    ax.text(bar.get_x() + bar.get_width() / 2, 
            label_pos, 
            f'{height:,.0f}', 
            ha='center', 
            va=va_pos, 
            color='black', 
            fontsize=7.5)
    

plt.savefig("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_month.png", dpi=100, bbox_inches="tight")



# %%
#Read data from Excel (your cleaned_financials.xlsx)
new_df = pd.read_excel("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/cleaned_financials.xlsx", sheet_name="Financials")

#Write summaries to Excel
with pd.ExcelWriter("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/fpa_model.xlsx", engine="openpyxl") as writer:
    new_df.to_excel(writer, sheet_name="Raw_Data", index=False)
    profit_by_year.to_excel(writer, sheet_name="Profit_by_Year", index=False)
    profit_by_country.to_excel(writer, sheet_name="Profit_by_Country", index=False)
    profit_by_month.to_excel(writer, sheet_name="Profit_by_Month", index=False)
    profit_by_product.to_excel(writer, sheet_name="Profit_by_Product", index=False)
    profit_by_segment.to_excel(writer, sheet_name="Profit_by_Segment", index=False)
    counts.to_excel(writer, sheet_name="Enterprise_Discount_Band", index=False)

#drop graphs into tab for readily use
book = load_workbook("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/fpa_model.xlsx")
sheet = book.create_sheet("Charts")
img1 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_year.png")
img2 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_country.png")
img3 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_month.png")
img4 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_product.png")
img5 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/profit_by_segment.png")
img6 = Image("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/Graph_Visuals/enterprise_discount_frequency.png")
sheet.add_image(img1, "A1")
sheet.add_image(img2, "A20")
sheet.add_image(img3, "A40")
sheet.add_image(img4, "A60")
sheet.add_image(img5, "A80")
sheet.add_image(img6, "A100")
book.save("C:/Users/Chris/Desktop/Documents/Company Financials Model Project/fpa_model.xlsx")
print("FP&A model updated in Excel!")

# %% [markdown]
# 

# %% [markdown]
# 


